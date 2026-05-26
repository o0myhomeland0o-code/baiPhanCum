import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# 1. ĐỌC VÀ LÀM SẠCH DỮ LIỆU GỐC
FILE_INPUT = 'diem.xlsx'
df_raw = pd.read_excel(FILE_INPUT, header=None)

# Lấy MSSV, Họ tên và bảng điểm từ hàng/cột tương ứng
mssv_list = df_raw.iloc[1, 3:].tolist()
name_list = df_raw.iloc[2, 3:].tolist()
grade_data = df_raw.iloc[4:, 3:].copy()

def chuan_hoa_diem(v):
    if pd.isna(v): return np.nan
    s = str(v).strip().lower().replace('nan', '')
    if s in ['-', '—', 'x', '']: return np.nan
    try:
        f = float(s)
        return f if 0 <= f <= 4.0 else np.nan
    except: return np.nan

clean_grades = grade_data.copy()
for col in clean_grades.columns:
    clean_grades[col] = clean_grades[col].apply(chuan_hoa_diem)

# Tính điểm GPA trung bình cho mỗi sinh viên
gpa_series = clean_grades.mean(axis=0)

# 2. ĐÁNH GIÁ CHẤT LƯỢNG CHI TIẾT TỪNG MÔN HỌC
def phan_tich_mon_hoc(col_data):
    valid_grades = col_data.dropna()
    total_subjects = len(valid_grades)
    if total_subjects == 0:
        return "Không có dữ liệu điểm"
    
    # Tính toán số môn xuất sắc (A từ 3.5 trở lên) và môn nợ học lại (F dưới 1.0)
    num_A = sum(valid_grades >= 3.5)
    num_F = sum(valid_grades < 1.0)
    
    if num_F > 0:
        return f"Còn nợ {int(num_F)} môn (F)"
    elif num_A >= (total_subjects * 0.4):
        return f"Vượt trội ({int(num_A)} môn đạt điểm A)"
    else:
        return "Ổn định"

# Tiến hành phân tích chi tiết môn học cho từng sinh viên
danh_gia_list = []
for col_name in clean_grades.columns:
    danh_gia_list.append(phan_tich_mon_hoc(clean_grades[col_name]))

# Gom dữ liệu hoàn chỉnh vào bảng dữ liệu chính
student_df = pd.DataFrame({
    'MSSV': mssv_list, 
    'Ho_va_ten': name_list, 
    'GPA': gpa_series.values,
    'Danh_gia': danh_gia_list
})
student_df = student_df.dropna(subset=['GPA']).reset_index(drop=True)

def phan_loai(gpa):
    if gpa >= 3.2: return 'Giỏi'
    elif gpa >= 2.5: return 'Khá'
    else: return 'Trung bình'

student_df['Xep_loai'] = student_df['GPA'].apply(phan_loai)
student_df['GPA'] = student_df['GPA'].round(2)

# Sắp xếp danh sách toàn bộ lớp học theo thứ tự điểm GPA giảm dần
df_sorted = student_df.sort_values(by='GPA', ascending=False).reset_index(drop=True)

# Chia tách dữ liệu theo từng nhóm đối tượng riêng để xử lý xuất khối cột
df_gioi = df_sorted[df_sorted['Xep_loai'] == 'Giỏi'].reset_index(drop=True)
df_kha = df_sorted[df_sorted['Xep_loai'] == 'Khá'].reset_index(drop=True)
df_tb = df_sorted[df_sorted['Xep_loai'] == 'Trung bình'].reset_index(drop=True)
total_count = len(df_sorted)

# ── 3. TỰ ĐỘNG VẼ BIỂU ĐỒ BẰNG MATPLOTLIB VÀ LƯU THÀNH ẢNH ──
counts = [len(df_gioi), len(df_kha), len(df_tb)]

labels = [
    'Giỏi (>=3.2)',
    'Khá (2.5-3.2)',
    'Trung bình (<2.5)'
]

colors = ['#5B9BD5', '#ED7D31', '#70AD47']

fig, ax = plt.subplots(figsize=(10, 6))

bars = ax.bar(
    labels,
    counts,
    color=colors,
    edgecolor='#D9D9D9',
    width=0.45
)

# Tiêu đề
ax.set_title(
    'Biểu đồ thống kê số lượng SV theo nhóm đối tượng',
    fontsize=18,
    fontweight='bold',
    pad=20
)

# Tên trục Y
ax.set_ylabel(
    'Số lượng sinh viên',
    fontsize=14
)

# Grid ngang
ax.grid(axis='y', linestyle='--', alpha=0.5)

# Hiện số lượng + phần trăm trên cột
for bar in bars:
    yval = bar.get_height()

    ax.text(
        bar.get_x() + bar.get_width()/2,
        yval + 0.5,
        f"{int(yval)} SV\n({yval/total_count*100:.1f}%)",
        ha='center',
        va='bottom',
        fontsize=12,
        fontweight='bold'
    )

# Ẩn viền trên và phải
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

plt.tight_layout()

# Lưu ảnh
chart_path = 'bieu_do_.png'
plt.savefig(chart_path, dpi=150)

plt.show()
plt.close()

# 4. Xuất EXCEL và chia thành 3 phần
wb = Workbook()
COLOR_MAP = {'Giỏi': '5B9BD5', 'Khá': 'ED7D31', 'Trung bình': '70AD47'}
HEADER_FILL = PatternFill('solid', start_color='2F5597', end_color='2F5597')
HEADER_FONT = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Segoe UI', size=15, bold=True, color='1F4E78')
TEXT_FONT = Font(name='Segoe UI', size=10, bold=False, color='000000')
BOLD_FONT = Font(name='Segoe UI', size=10, bold=True, color='000000')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

# P1:
ws1 = wb.active; ws1.title = 'Xếp hạng cả lớp'
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells('A1:F1')
ws1['A1'] = 'DANH SÁCH XẾP PHÂN LOẠI SINH VIÊN LỚP (GPA GIẢM DẦN)'
ws1['A1'].font = TITLE_FONT; ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

# Tiêu đề cột Đánh giá chi tiết môn học mới
headers = ['STT', 'MSSV', 'Họ và Tên', 'GPA Hệ 4', 'Xếp Loại', 'Đánh giá trình độ']
for col_idx, text in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col_idx, value=text)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[3].height = 25

for idx, row in df_sorted.iterrows():
    r_idx = idx + 4
    xl = row['Xep_loai']
    fill_color = PatternFill('solid', start_color=COLOR_MAP[xl], end_color=COLOR_MAP[xl])
    
    vals = [idx + 1, row['MSSV'], row['Ho_va_ten'], row['GPA'], xl, row['Danh_gia']]
    for c_idx, val in enumerate(vals, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=val)
        cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill_color
        cell.alignment = Alignment(horizontal='center' if c_idx in [1, 2, 4, 5] else 'left', vertical='center')
        if c_idx == 4: cell.number_format = '0.00'
    ws1.row_dimensions[r_idx].height = 20

widths1 = [6, 16, 28, 12, 14, 32]  # Cột đánh giá rộng rãi để tránh tràn chữ
for c_idx, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(c_idx)].width = w


#P2: TAB TỔNG HỢP THỐNG KÊ & BIỂU ĐỒ
ws2 = wb.create_sheet(title='Tổng hợp thống kê')
ws2.views.sheetView[0].showGridLines = True
ws2.merge_cells('A1:D1')
ws2['A1'] = 'BÁO CÁO THỐNG KÊ CHẤT LƯỢNG LỚP HỌC K58KTP'
ws2['A1'].font = TITLE_FONT; ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 40

headers_tk = ['Nhóm đối tượng', 'Tiêu chí phân loại', 'Số lượng SV', 'Tỷ lệ phần trăm']
for col_idx, text in enumerate(headers_tk, 1):
    cell = ws2.cell(row=3, column=col_idx, value=text)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[3].height = 25

tk_rows_data = [
    ('Giỏi', 'GPA >= 3.2', len(df_gioi), len(df_gioi)/total_count),
    ('Khá', '2.5 <= GPA < 3.2', len(df_kha), len(df_kha)/total_count),
    ('Trung bình', 'GPA < 2.5', len(df_tb), len(df_tb)/total_count)
]

for idx, (nhom, tchi, count, pct) in enumerate(tk_rows_data):
    r_idx = idx + 4
    fill_color = PatternFill('solid', start_color=COLOR_MAP[nhom], end_color=COLOR_MAP[nhom])
    c1 = ws2.cell(row=r_idx, column=1, value=nhom)
    c2 = ws2.cell(row=r_idx, column=2, value=tchi)
    c3 = ws2.cell(row=r_idx, column=3, value=count)
    c4 = ws2.cell(row=r_idx, column=4, value=pct)
    for cell in [c1, c2, c3, c4]:
        cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill_color
        cell.alignment = Alignment(horizontal='center', vertical='center')
    c4.number_format = '0.0%'
    ws2.row_dimensions[r_idx].height = 22

r_total = 7
ws2.cell(row=r_total, column=1, value='Tổng cộng lớp học')
ws2.cell(row=r_total, column=2, value='-')
ws2.cell(row=r_total, column=3, value=total_count)
ws2.cell(row=r_total, column=4, value=1.0)
for c_idx in range(1, 5):
    cell = ws2.cell(row=r_total, column=c_idx)
    cell.font = BOLD_FONT; cell.border = thin_border; cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=r_total, column=4).number_format = '0.0%'
ws2.row_dimensions[r_total].height = 22

widths2 = [18, 22, 15, 18]
for c_idx, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(c_idx)].width = w

# Nhúng biểu đồ vào trang số 2
img = OpenpyxlImage(chart_path)
ws2.add_image(img, 'A10')


# P3: TAB CHIA CỘT THEO NHÓM
ws3 = wb.create_sheet(title='Chia cột theo nhóm')
ws3.views.sheetView[0].showGridLines = True
ws3.merge_cells('A1:N1')
ws3['A1'] = 'DANH SÁCH THEO DÕI SONG SONG THEO KHỐI ĐỐI TƯỢNG (KÈM ĐÁNH GIÁ MÔN)'
ws3['A1'].font = TITLE_FONT; ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

def thiet_lap_cum_cot(ws, start_col, ten_nhom, mã_màu_nền):
    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col+3)
    top_cell = ws.cell(row=3, column=start_col, value=ten_nhom)
    top_cell.font = Font(name='Segoe UI', size=11, bold=True, color='1F4E78')
    top_cell.fill = PatternFill('solid', start_color=mã_màu_nền, end_color=mã_màu_nền)
    top_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    tieu_de_con = ['STT', 'Họ và Tên', 'GPA', 'Đánh giá môn']
    for i, txt in enumerate(tieu_de_con):
        cell = ws.cell(row=4, column=start_col+i, value=txt)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

thiet_lap_cum_cot(ws3, start_col=1, ten_nhom='🏆 NHÓM 1: GIỎI (GPA >= 3.2)', mã_màu_nền='E2EFDA')
thiet_lap_cum_cot(ws3, start_col=6, ten_nhom='✅ NHÓM 2: KHÁ (2.5 <= GPA < 3.2)', mã_màu_nền='FFF2CC')
thiet_lap_cum_cot(ws3, start_col=11, ten_nhom='⚠️ NHÓM 3: TRUNG BÌNH (GPA < 2.5)', mã_màu_nền='FCE4D6')

ws3.row_dimensions[3].height = 25
ws3.row_dimensions[4].height = 22

max_rows = max(len(df_gioi), len(df_kha), len(df_tb))
for idx in range(max_rows):
    r_idx = idx + 5
    ws3.row_dimensions[r_idx].height = 20
    
    # Khối Giỏi
    if idx < len(df_gioi):
        row = df_gioi.iloc[idx]
        fill = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
        for offset, val in enumerate([idx+1, row['Ho_va_ten'], row['GPA'], row['Danh_gia']]):
            cell = ws3.cell(row=r_idx, column=1+offset, value=val)
            cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if offset in [0, 2] else 'left', vertical='center')
            if offset == 2: cell.number_format = '0.00'
            
    # Khối Khá
    if idx < len(df_kha):
        row = df_kha.iloc[idx]
        fill = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')
        for offset, val in enumerate([idx+1, row['Ho_va_ten'], row['GPA'], row['Danh_gia']]):
            cell = ws3.cell(row=r_idx, column=6+offset, value=val)
            cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if offset in [0, 2] else 'left', vertical='center')
            if offset == 2: cell.number_format = '0.00'
            
    # Khối Trung bình
    if idx < len(df_tb):
        row = df_tb.iloc[idx]
        fill = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
        for offset, val in enumerate([idx+1, row['Ho_va_ten'], row['GPA'], row['Danh_gia']]):
            cell = ws3.cell(row=r_idx, column=11+offset, value=val)
            cell.font = TEXT_FONT; cell.border = thin_border; cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if offset in [0, 2] else 'left', vertical='center')
            if offset == 2: cell.number_format = '0.00'

col_widths_s3 = {
    1: 5, 2: 24, 3: 8, 4: 26,    # Cụm Giỏi
    5: 4,                        # Phân cách
    6: 5, 7: 24, 8: 8, 9: 26,    # Cụm Khá
    10: 4,                       # Phân cách
    11: 5, 12: 24, 13: 8, 14: 26 # Cụm Trung bình
}
for col_num, width in col_widths_s3.items():
    ws3.column_dimensions[get_column_letter(col_num)].width = width

# 5. LƯU FILE KẾT QUẢ
OUTPUT_PATH = 'Phan_cum_lop_K58KTP.xlsx'
wb.save(OUTPUT_PATH)
print("=" * 70)
print(f" Sửa code.py thành công! File Excel chia 3 tab riêng biệt: ")
print(f" {OUTPUT_PATH}")
print("=" * 70)